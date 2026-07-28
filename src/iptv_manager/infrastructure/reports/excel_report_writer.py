"""Excel report writer (.xlsx) using openpyxl.

Produces a multi-sheet workbook - Summary, Streams, Logos, Duplicates,
EPG Issues - one sheet per concern, so a playlist maintainer can jump
straight to what they care about instead of scrolling one giant flat
table. Sheets are only added for data that's actually available in the
report (e.g. no "Streams" sheet if stream validation wasn't run).
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from iptv_manager.application.dto.validation_report import ValidationReport

_HEADER_FONT = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=14)


class ExcelReportWriter:
    def write(self, report: ValidationReport, path: Path) -> None:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "Summary"
        self._write_summary_sheet(summary_sheet, report)

        if report.stream_summary is not None:
            self._write_streams_sheet(workbook.create_sheet("Streams"), report)
        if report.logo_summary is not None:
            self._write_logos_sheet(workbook.create_sheet("Logos"), report)
        if report.merge_result is not None and report.merge_result.duplicate_urls:
            self._write_duplicates_sheet(workbook.create_sheet("Duplicates"), report)
        if report.epg_comparison is not None:
            self._write_epg_sheet(workbook.create_sheet("EPG Issues"), report)

        workbook.save(path)

    def _write_summary_sheet(self, ws: Worksheet, report: ValidationReport) -> None:
        ws.append(["IPTV Playlist Validation Report"])
        ws["A1"].font = _TITLE_FONT
        ws.append(["Generated at", report.generated_at.isoformat()])
        ws.append(["Master playlist", report.master_playlist_name])
        ws.append([])

        if report.merge_result is not None:
            mr = report.merge_result
            self._section_header(ws, "Merge summary")
            ws.append(["Channels before dedup", mr.total_channels_before])
            ws.append(["Channels after dedup", mr.total_channels_after])
            ws.append(["Duplicate URLs removed", mr.removed_duplicate_url_count])
            ws.append(["Duplicate tvg-id groups", len(mr.duplicate_tvg_ids)])
            ws.append([])

        if report.stream_summary is not None:
            summary = report.stream_summary
            self._section_header(ws, "Stream validation summary")
            ws.append(["Total streams", summary.total])
            ws.append(["Online", summary.online_count])
            ws.append(["Offline", summary.offline_count])
            ws.append([])

        if report.logo_summary is not None:
            logos = report.logo_summary
            self._section_header(ws, "Logo validation summary")
            ws.append(["Total channels", logos.total])
            ws.append(["Reachable", logos.reachable_count])
            ws.append(["Missing", logos.missing_count])
            ws.append(["Unreachable", logos.unreachable_count])
            ws.append([])

        if report.epg_comparison is not None:
            epg = report.epg_comparison
            self._section_header(ws, "EPG comparison summary")
            ws.append(["Missing tvg-id", len(epg.missing_tvg_id)])
            ws.append(["Invalid tvg-id", len(epg.invalid_tvg_id)])
            ws.append(["Duplicate tvg-id", len(epg.duplicate_tvg_id)])
            ws.append(["Unused EPG entries", len(epg.unused_epg_entries)])

    def _section_header(self, ws: Worksheet, title: str) -> None:
        ws.append([title])
        ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT

    def _header_row(self, ws: Worksheet, headers: list[str]) -> None:
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = _HEADER_FONT

    def _write_streams_sheet(self, ws: Worksheet, report: ValidationReport) -> None:
        self._header_row(
            ws, ["Name", "Group", "URL", "Status", "HTTP Status", "Response Time (ms)", "Error"]
        )
        for result in report.stream_summary.results:
            ws.append(
                [
                    result.channel.name,
                    str(result.channel.group_title),
                    str(result.channel.url),
                    result.status.value,
                    result.http_status,
                    result.response_time_ms,
                    result.error_message,
                ]
            )

    def _write_logos_sheet(self, ws: Worksheet, report: ValidationReport) -> None:
        self._header_row(ws, ["Name", "Logo URL", "Reachable", "HTTP Status", "Error"])
        for result in report.logo_summary.results:
            ws.append(
                [
                    result.channel.name,
                    result.channel.logo_url,
                    result.reachable,
                    result.http_status,
                    result.error_message,
                ]
            )

    def _write_duplicates_sheet(self, ws: Worksheet, report: ValidationReport) -> None:
        self._header_row(ws, ["Duplicate URL", "Kept Channel", "Removed Channel"])
        for group in report.merge_result.duplicate_urls:
            for removed in group.removed:
                ws.append([group.key, group.kept.name, removed.name])

    def _write_epg_sheet(self, ws: Worksheet, report: ValidationReport) -> None:
        epg = report.epg_comparison
        self._header_row(ws, ["Issue Type", "Channel / EPG ID", "Detail"])
        for channel in epg.missing_tvg_id:
            ws.append(["Missing tvg-id", channel.name, ""])
        for channel in epg.invalid_tvg_id:
            ws.append(["Invalid tvg-id", channel.name, str(channel.tvg_id)])
        for group in epg.duplicate_tvg_id:
            ws.append(
                ["Duplicate tvg-id", group.tvg_id, ", ".join(c.name for c in group.channels)]
            )
        for epg_channel in epg.unused_epg_entries:
            ws.append(["Unused EPG entry", epg_channel.id, epg_channel.primary_display_name or ""])

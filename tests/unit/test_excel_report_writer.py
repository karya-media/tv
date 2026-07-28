"""Unit tests for infrastructure.reports.excel_report_writer."""

from pathlib import Path

from openpyxl import load_workbook

from iptv_manager.infrastructure.reports.excel_report_writer import ExcelReportWriter
from tests.unit.report_test_helpers import build_full_report, build_minimal_report


def test_full_report_creates_all_expected_sheets(tmp_path: Path):
    output = tmp_path / "report.xlsx"
    ExcelReportWriter().write(build_full_report(), output)
    wb = load_workbook(output)
    assert set(wb.sheetnames) == {"Summary", "Streams", "Logos", "Duplicates", "EPG Issues"}


def test_minimal_report_only_creates_summary_sheet(tmp_path: Path):
    output = tmp_path / "report.xlsx"
    ExcelReportWriter().write(build_minimal_report(), output)
    wb = load_workbook(output)
    assert wb.sheetnames == ["Summary"]


def test_summary_sheet_contains_merge_counts(tmp_path: Path):
    output = tmp_path / "report.xlsx"
    ExcelReportWriter().write(build_full_report(), output)
    wb = load_workbook(output)
    values = [row[1] for row in wb["Summary"].iter_rows(values_only=True) if row[0] == "Channels after dedup"]
    assert values == [2]


def test_streams_sheet_has_one_row_per_result_plus_header(tmp_path: Path):
    output = tmp_path / "report.xlsx"
    ExcelReportWriter().write(build_full_report(), output)
    wb = load_workbook(output)
    rows = list(wb["Streams"].iter_rows(values_only=True))
    assert rows[0][0] == "Name"  # header
    assert len(rows) == 3  # header + 2 channels


def test_duplicates_sheet_lists_kept_and_removed(tmp_path: Path):
    output = tmp_path / "report.xlsx"
    ExcelReportWriter().write(build_full_report(), output)
    wb = load_workbook(output)
    rows = list(wb["Duplicates"].iter_rows(values_only=True))
    assert rows[1] == ("http://x.com/espn.m3u8", "ESPN", "ESPN Dup")
